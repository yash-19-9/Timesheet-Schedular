"""
create_sample_excel.py
Run this script to generate sample Excel upload files for the system.
Usage: python create_sample_excel.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def style_header(ws, header_row, fill_color="1e3a5f"):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    border = Border(
        bottom=Side(style='medium', color="4CAF50"),
        right=Side(style='thin', color="AAAAAA")
    )
    for col_num, header in enumerate(header_row, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = max(len(str(header)) + 6, 18)

    ws.row_dimensions[1].height = 24


# ── 1. Faculty Excel ─────────────────────────────────────────────────────────
wb_faculty = Workbook()
ws = wb_faculty.active
ws.title = "Faculty Data"

headers = ['name', 'subjects', 'max_hours', 'employee_id', 'department', 'email']
style_header(ws, headers)

faculty_rows = [
    ['Dr. Anjali Sharma',  'Mathematics,Statistics',           18, 'F001', 'Science',     'anjali.sharma@college.edu'],
    ['Prof. Ravi Kumar',   'Physics,Engineering Mechanics',    20, 'F002', 'Engineering', 'ravi.kumar@college.edu'],
    ['Dr. Priya Menon',    'Computer Science,Data Structures', 16, 'F003', 'IT',          'priya.menon@college.edu'],
    ['Mr. Suresh Patil',   'Chemistry,Environmental Science',  20, 'F004', 'Science',     'suresh.patil@college.edu'],
    ['Dr. Fatima Sheikh',  'English,Communication Skills',     14, 'F005', 'Humanities',  'fatima.sheikh@college.edu'],
    ['Prof. Arjun Nair',   'Electronics,VLSI Design',          18, 'F006', 'Engineering', 'arjun.nair@college.edu'],
]

alt_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
for i, row in enumerate(faculty_rows, 2):
    for col, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=col, value=val)
        if i % 2 == 0:
            cell.fill = alt_fill

wb_faculty.save('sample_excel/faculty_upload.xlsx')
print("✅ faculty_upload.xlsx created")

# ── 2. Courses Excel ─────────────────────────────────────────────────────────
wb_courses = Workbook()
ws = wb_courses.active
ws.title = "Courses Data"

headers = ['course_name', 'hours_required', 'course_code', 'department', 'semester']
style_header(ws, headers)

course_rows = [
    ['Mathematics',      4, 'MATH101', 'Science',     1],
    ['Physics',          3, 'PHY101',  'Engineering', 1],
    ['Computer Science', 4, 'CS101',   'IT',          1],
    ['Data Structures',  4, 'CS201',   'IT',          2],
    ['Chemistry',        3, 'CHEM101', 'Science',     1],
    ['English',          2, 'ENG101',  'Humanities',  1],
    ['Statistics',       3, 'STAT101', 'Science',     2],
    ['Electronics',      4, 'ELE201',  'Engineering', 2],
]

for i, row in enumerate(course_rows, 2):
    for col, val in enumerate(row, 1):
        cell = wb_courses.active.cell(row=i, column=col, value=val)
        if i % 2 == 0:
            cell.fill = alt_fill

wb_courses.save('sample_excel/courses_upload.xlsx')
print("✅ courses_upload.xlsx created")

# ── 3. Classrooms Excel ───────────────────────────────────────────────────────
wb_rooms = Workbook()
ws = wb_rooms.active
ws.title = "Classrooms Data"

headers = ['room_name', 'capacity', 'building', 'room_type']
style_header(ws, headers)

room_rows = [
    ['Room 101',  60, 'Block A', 'lecture'],
    ['Room 102',  60, 'Block A', 'lecture'],
    ['Room 201',  80, 'Block B', 'lecture'],
    ['Lab 001',   30, 'Block C', 'lab'],
    ['Lab 002',   30, 'Block C', 'lab'],
    ['Seminar A', 40, 'Block D', 'seminar'],
]

for i, row in enumerate(room_rows, 2):
    for col, val in enumerate(row, 1):
        cell = wb_rooms.active.cell(row=i, column=col, value=val)
        if i % 2 == 0:
            cell.fill = alt_fill

wb_rooms.save('sample_excel/classrooms_upload.xlsx')
print("✅ classrooms_upload.xlsx created")

print("\n📁 All sample Excel files saved to: sample_excel/")
print("   Use these files in the Upload Data page of the app.")

if __name__ == '__main__':
    import os
    os.makedirs('sample_excel', exist_ok=True)
    print("Creating sample Excel templates...")
