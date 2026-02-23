"""
excel_upload.py - Excel parsing & bulk import logic
AI-Based Intelligent Faculty Workload and Scheduling System

Parses three Excel templates:
  1. faculty_upload.xlsx   → Faculty model
  2. courses_upload.xlsx   → Course model
  3. classrooms_upload.xlsx → Classroom model
"""

import logging
from openpyxl import load_workbook
from django.contrib.auth.models import User
from .models import Faculty, Course, Classroom

logger = logging.getLogger(__name__)


def parse_faculty_excel(file_obj):
    """
    Parse faculty Excel file.
    Expected columns: name | subjects | max_hours | employee_id | department | email

    Returns (created_count, skipped_count, errors[])
    """
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active

    created, skipped, errors = 0, 0, []
    headers = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            headers = [str(h).strip().lower() if h else '' for h in row]
            continue

        if all(cell is None for cell in row):
            continue  # skip blank rows

        try:
            data = dict(zip(headers, row))

            name = str(data.get('name', '')).strip()
            subjects = str(data.get('subjects', '')).strip()
            max_hours = int(data.get('max_hours', 20) or 20)
            employee_id = str(data.get('employee_id', '')).strip() or None
            department = str(data.get('department', '')).strip()
            email = str(data.get('email', '')).strip()

            if not name:
                errors.append(f"Row {row_idx}: 'name' is required.")
                skipped += 1
                continue

            if not subjects:
                errors.append(f"Row {row_idx}: 'subjects' is required for {name}.")
                skipped += 1
                continue

            # Create/update Django user for faculty login
            username = name.lower().replace(' ', '_')
            user, _ = User.objects.get_or_create(username=username)
            user.set_password('faculty123')  # default password
            user.first_name = name.split()[0] if name else ''
            user.last_name = ' '.join(name.split()[1:]) if len(name.split()) > 1 else ''
            user.email = email
            user.save()

            # Create/update faculty
            Faculty.objects.update_or_create(
                name=name,
                defaults={
                    'user': user,
                    'subjects': subjects,
                    'max_hours': max_hours,
                    'employee_id': employee_id,
                    'department': department,
                    'email': email,
                    'is_active': True,
                }
            )
            created += 1

        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
            skipped += 1

    wb.close()
    return created, skipped, errors


def parse_courses_excel(file_obj):
    """
    Parse courses Excel file.
    Expected columns: course_name | hours_required | course_code | department | semester

    Returns (created_count, skipped_count, errors[])
    """
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active

    created, skipped, errors = 0, 0, []
    headers = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            headers = [str(h).strip().lower() if h else '' for h in row]
            continue

        if all(cell is None for cell in row):
            continue

        try:
            data = dict(zip(headers, row))

            course_name = str(data.get('course_name', '')).strip()
            hours_required = int(data.get('hours_required', 4) or 4)
            course_code = str(data.get('course_code', '')).strip()
            department = str(data.get('department', '')).strip()
            semester = int(data.get('semester', 1) or 1)

            if not course_name:
                errors.append(f"Row {row_idx}: 'course_name' is required.")
                skipped += 1
                continue

            Course.objects.update_or_create(
                course_name=course_name,
                defaults={
                    'hours_required': hours_required,
                    'course_code': course_code,
                    'department': department,
                    'semester': semester,
                    'is_active': True,
                }
            )
            created += 1

        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
            skipped += 1

    wb.close()
    return created, skipped, errors


def parse_classrooms_excel(file_obj):
    """
    Parse classrooms Excel file.
    Expected columns: room_name | capacity | building | room_type

    Returns (created_count, skipped_count, errors[])
    """
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active

    created, skipped, errors = 0, 0, []
    headers = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            headers = [str(h).strip().lower() if h else '' for h in row]
            continue

        if all(cell is None for cell in row):
            continue

        try:
            data = dict(zip(headers, row))

            room_name = str(data.get('room_name', '')).strip()
            capacity = int(data.get('capacity', 60) or 60)
            building = str(data.get('building', '')).strip()
            room_type = str(data.get('room_type', 'lecture')).strip().lower()

            valid_types = ['lecture', 'lab', 'seminar', 'tutorial']
            if room_type not in valid_types:
                room_type = 'lecture'

            if not room_name:
                errors.append(f"Row {row_idx}: 'room_name' is required.")
                skipped += 1
                continue

            Classroom.objects.update_or_create(
                room_name=room_name,
                defaults={
                    'capacity': capacity,
                    'building': building,
                    'room_type': room_type,
                    'is_available': True,
                }
            )
            created += 1

        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
            skipped += 1

    wb.close()
    return created, skipped, errors
